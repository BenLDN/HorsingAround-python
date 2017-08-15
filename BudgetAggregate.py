#Digital Wallets 2018 Budget Aggregator
#Only works with the official budget templates

from openpyxl import *
from os import walk

path = r'C:\Users\bencekulcsar\Desktop\XLS\\' #path of the target folder where the budget templates are located - can be changed if necessary

sum_counter_opex = sum_counter_hc = 1 #row pointers for the aggregator file
first_file=True #needed b/c we only copy the header from the first file & not from the rest
wb_sum = Workbook() #workbook to store the aggregated data
ws_sum_opex = wb_sum.active #the only sheet after creating the wb will be the opex aggregator sheet
ws_sum_opex.title="OPEX_Aggregate"
ws_sum_hc=wb_sum.create_sheet("HC_Aggregate") #create a new sheet for headcount aggregation

#file_list stores the list of files (templates) from the target folder. If the consolidation / aggreation file already exists in the target folder, it's excluded from this list

file_list=[]

for (dp, dn, fn) in walk(path):
    file_list.extend(fn)
    break
try:
    file_list.remove('BudgetSummary.xlsx')
except ValueError:
    pass

total_files=len(file_list) #total number of templates in target folder
done=0 #number of files already processed

#looping throught the templates

for file in file_list:
    wb_tmp=load_workbook(path + file, data_only=True)
    ws_tmp_opex=wb_tmp["OPEX Input"]
    ws_tmp_hc=wb_tmp["Headcount Input"]

    start_row_opex=4
    start_row_hc=3

    last_line_opex=wb_tmp["Dim"]["s2"].value+3 #last line of the OPEX Input sheet is stored in cell S2 of the Dim sheet (COUNTA function)
    last_line_hc=wb_tmp["Dim"]["s3"].value+2 #last line of the Headcount Input sheet is stored in cell S3 of the Dim sheet (COUNTA function)

    #we only copy the header from the first file. After this first_file is set to False and the start_row variables are 3,4 instead of 2,3
    
    if first_file==True:
        first_file=False
        start_row_opex=3
        start_row_hc=2
        
    #looping through the lines of the OPEX Input sheet and copying ---RELEVANT--- cells to the aggregator file
    
    for line in range(start_row_opex, last_line_opex+1):
    
        ws_sum_opex.cell(row=sum_counter_opex,column=1).value=ws_tmp_opex.cell(row=line,column=3).value
        ws_sum_opex.cell(row=sum_counter_opex,column=2).value=ws_tmp_opex.cell(row=line,column=5).value
        ws_sum_opex.cell(row=sum_counter_opex,column=3).value=ws_tmp_opex.cell(row=line,column=6).value
        ws_sum_opex.cell(row=sum_counter_opex,column=4).value=ws_tmp_opex.cell(row=line,column=7).value
        ws_sum_opex.cell(row=sum_counter_opex,column=5).value=ws_tmp_opex.cell(row=line,column=8).value

        ws_sum_opex.cell(row=sum_counter_opex,column=6).value=ws_tmp_opex.cell(row=line,column=16).value
        ws_sum_opex.cell(row=sum_counter_opex,column=7).value=ws_tmp_opex.cell(row=line,column=17).value
        ws_sum_opex.cell(row=sum_counter_opex,column=8).value=ws_tmp_opex.cell(row=line,column=18).value
        ws_sum_opex.cell(row=sum_counter_opex,column=9).value=ws_tmp_opex.cell(row=line,column=19).value
        ws_sum_opex.cell(row=sum_counter_opex,column=10).value=ws_tmp_opex.cell(row=line,column=20).value
        ws_sum_opex.cell(row=sum_counter_opex,column=11).value=ws_tmp_opex.cell(row=line,column=21).value
        ws_sum_opex.cell(row=sum_counter_opex,column=12).value=ws_tmp_opex.cell(row=line,column=22).value
        ws_sum_opex.cell(row=sum_counter_opex,column=13).value=ws_tmp_opex.cell(row=line,column=23).value
        ws_sum_opex.cell(row=sum_counter_opex,column=14).value=ws_tmp_opex.cell(row=line,column=24).value
        ws_sum_opex.cell(row=sum_counter_opex,column=15).value=ws_tmp_opex.cell(row=line,column=25).value
        ws_sum_opex.cell(row=sum_counter_opex,column=16).value=ws_tmp_opex.cell(row=line,column=26).value
        ws_sum_opex.cell(row=sum_counter_opex,column=17).value=ws_tmp_opex.cell(row=line,column=27).value
        
        sum_counter_opex+=1

    #looping through the lines of the Headcount Input sheet and copying ---ALL--- cells to the aggregator file

    for line in range(start_row_hc, 20):
        for col_loop in range(1,22): #as all cells in the line are to be copied it's shorter to do this with a nested loop for cells

            ws_sum_hc.cell(row=sum_counter_hc,column=col_loop).value=ws_tmp_hc.cell(row=line,column=col_loop).value

        sum_counter_hc+=1

    done+=1

    print(file + " (" + str(last_line_opex-3) + " OPEX line(s)) and " + str(last_line_hc-2) + " headcount line(s) processed.")
    print(str(done) + " of " + str(total_files) + " done.")
    
wb_sum.save(path + "BudgetSummary.xlsx") #save aggregator file in target folder


