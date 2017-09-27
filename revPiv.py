from openpyxl import *

wb=load_workbook("reversePivot.xlsx")
ws=wb["working"]

mat=[]

col_attrs=3
row_attrs=2

last_row=ws.max_row
last_col=ws.max_column

for c in range(col_attrs+1,last_col+1):
    for r in range(row_attrs+1,last_row+1):

        app_line=[]

        for col_attr in range(1,col_attrs+1):
            app_line.append(ws.cell(row=r, column=col_attr).value)
        for row_attr in range(1,row_attrs+1):
            app_line.append(ws.cell(row=row_attr, column=c).value)
        
        app_line.append(ws.cell(row=r, column=c).value)
        
        mat.append(app_line)
        
for row in mat:
    for cell in row:
        print(cell, end=' ')
    print("")

    c2=0
r2=0

wb_new = Workbook()
ws_new=wb_new.active
  
for row in mat:
    r2+=1
    c2=0
    for cell in row:
        c2+=1
        ws_new.cell(row=r2, column=c2).value=cell


wb_new.save("general.xlsx")
